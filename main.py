from itertools import product
from generate_RBN import generate_RBN
from find_2point_attractor_network import find_two_point_attractor
from measure_basin_difference import measure_basin_difference
from measure_barrier_to_noise import measure_barrier_to_noise
import pandas as pd
import openpyxl

def main(initial=None):
    
    # takes random number of nodes
    node_number = int(input("Enter the number of nodes: "))

    # takes random average degree
    degree_k = int(input("Enter the average degree k: "))

    """
    # store the random Boolean nework if the networks are 2-point attractors
    two_points_attractors_networks = []
    while len(two_points_attractors_networks) < 100:
        #generate random Boolean network
        rbn = generate_RBN(node_number, degree_k)
        two_or_not, ID_state_combination = find_two_point_attractor(rbn)
        if two_or_not:
            two_points_attractors_networks.append(rbn)

    # measure and record the basin difference and barrier to noise
    basin_diff_barrier_to_noise = []
    for i in range (0, len(two_points_attractors_networks)):
        two_or_not, ID_state_combination = find_two_point_attractor(two_points_attractors_networks[i])
        basin_diff = measure_basin_difference(ID_state_combination)
        barrier_to_noise = measure_barrier_to_noise(two_points_attractors_networks[i])
        basin_diff_barrier_to_noise.append((basin_diff, barrier_to_noise))

     # generate the data frame to change it into excel
    df = pd.DataFrame(basin_diff_barrier_to_noise, columns = ['basin_difference', 'barrier to noise'])
    with pd.ExcelWriter('basin_diff_barrier_to_noise.xlsx') as writer:
        df.to_excel(writer, sheet_name='sheet1')
    """
    #example of generating data frame
#    df = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
#                  index=['one', 'two', 'three'], columns=['a', 'b', 'c'])

""""
    data_size = 0
    while data_size < 100:
    
        #generate two point attractor network
        two_points_attractors_networks = []
        while len(two_points_attractors_networks) < 100:
            #generate random Boolean network
            rbn = generate_RBN(node_number, degree_k)
            two_or_not, ID_state_combination = find_two_point_attractor(rbn)
            if two_or_not:
                two_points_attractors_networks.append(rbn)

        # measure and record the basin difference and barrier to noise
        basin_diff_barrier_to_noise = []
        for i in range (0, len(two_points_attractors_networks)):
            two_or_not, ID_state_combination = find_two_point_attractor(two_points_attractors_networks[i])
            basin_diff = measure_basin_difference(ID_state_combination)
            barrier_to_noise = measure_barrier_to_noise(two_points_attractors_networks[i])
            basin_diff_barrier_to_noise.append((basin_diff, barrier_to_noise))
        
        if data_size == 0:
            # Open the existing workbook or create a new one
            try:
                workbook = openpyxl.load_workbook('data.xlsx')
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            # Remove the default sheet created by Workbook
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)
            
        # Generate a unique sheet name or index
        sheet_name = f"Sheet{len(workbook.sheetnames) + 1}"

        # Create a new sheet with the generated name
        sheet = workbook.create_sheet(title=sheet_name)

        # Write data to the sheet
        for row in basin_diff_barrier_to_noise:
            sheet.append(row)

        # Save the workbook
        workbook.save('data.xlsx')

        data_size += 1

    print(basin_diff_barrier_to_noise)

"""
    # generate the list to make the data frame that is compatible to yEd
    rbn = generate_RBN(node_number, degree_k)

    #generate the list to make the data frame that is compatible to yEd
    state_trans_data_frame = []
    for lst in rbn:
        for j in range(len(lst) - 1):
            list_1 = lst[j]
            list_2 = lst[j + 1]
            list_3 = [list_1, list_2]
            state_trans_data_frame.append(list_3)
    
    # filter the redandant element in state_trans_data_frame
    filtered_data_frame = []
    for sublist in state_trans_data_frame:
        if sublist not in filtered_data_frame:
            filtered_data_frame.append(sublist)

    # generate the data frame to chnage it into excel
    df = pd.DataFrame(filtered_data_frame)
    #example of generating data frame
#    df = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
#                  index=['one', 'two', 'three'], columns=['a', 'b', 'c'])

    # generate the excel file with contents in data frame
    with pd.ExcelWriter('pandas_to_excel.xlsx') as writer:
        df.to_excel(writer, sheet_name='sheet1')

    # find an attractor from one random initial node
 #   attractor= update_list(initial, bool_func, degree_k)
 #   print (attractor)

if __name__ == "__main__":
    
    main()
