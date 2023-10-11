from generate_noise_trajectory import generate_noise_trajectory
#from connect_nonredundant_similar_state import connect_similar_state
from connect_similar_state import connect_similar_state
from measure_degree_distribution import measure_degree_distribution
import openpyxl


def main():
    # takes random number of nodes
    node_number = int(input("Enter the number of nodes: "))

    # takes random average degree
    degree_k = int(input("Enter the average degree k: "))

    data_size = 0
    while data_size < 2:
        # generate noise trajectory
        noise_trajectory = generate_noise_trajectory(node_number, degree_k)

        # connect similar state in the noise trajectory
        connection = connect_similar_state(noise_trajectory)
        if connection:
            # measure degree distribution of noise trajectory
            degree_distribution_dict = measure_degree_distribution(connection)

            # convert the dictionary into list in order to save the data using openpyxl
            degree_distribution_list = [(key, value) for key, value in degree_distribution_dict.items() if value != 0]

            if data_size == 0:
                # Open the existing workbook or create a new one
                try:
                    workbook = openpyxl.load_workbook('degree_distribution_data.xlsx')
                except FileNotFoundError:
                    workbook = openpyxl.Workbook()
    
                # Remove the default sheet created by Workbook
                default_sheet = workbook['Sheet']
                workbook.remove(default_sheet)
    
            # Generate a unique sheet name or index
            sheet_name = f"Sheet{len(workbook.sheetnames) + 1}"
    
            # Create a new sheet with the generated name
            sheet = workbook.create_sheet(title=sheet_name)
    
            # Write data to the sheet
            for row in degree_distribution_list:
                sheet.append(row)

            # Write data to the sheet
            # Write column headers
            sheet['A1'] = 'number_of_links_k'
            sheet['B1'] = 'number_of_states_with_k_links'
    
            # Write data rows
            for row_idx, row in enumerate(degree_distribution_list, start=2):  # Start from row 2 (index 1)
                key, value = row
                sheet.cell(row=row_idx, column=1, value=key)
                sheet.cell(row=row_idx, column=2, value=value)
            
            # Save the workbook
            workbook.save('degree_distribution_data.xlsx')

            data_size += 1

        else:
            print("No similar states in noise trajectory")
            break


if __name__ == "__main__":
    main()
